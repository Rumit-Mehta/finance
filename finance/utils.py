import finance.constants as con
from openpyxl import load_workbook
from datetime import timedelta


EXCEL_FILE = con.EXCEL_FILE
SHEET_NAME = "Transactions"

# get the date for the last entry for a spefific account eg. monzo, amex etc
def latest_entry(account):

    # load the workbook and sheet
    workbook = load_workbook(EXCEL_FILE)
    ws = workbook['Tracking']

    # get the indecies for the date and account column
    date_column = 2
    account_column = 8

    #reverse iterate through the sheet
    latest_date = None
    for row in reversed(list(ws.iter_rows(min_row=12, values_only=True))):  # Start from row 12
        date_value, account_value = row[date_column], row[account_column]  # Adjust based on actual columns
        if account_value and account in str(account_value):
            latest_date = date_value
            break
    
    # Print the latest date found
    if latest_date:
        print("Latest Monzo transaction date:", latest_date)
        print("Day after Monzo transaction date:", latest_date + timedelta(days=1))
        # you are more interested in the start date which is 1 more than the latest date. 
        latest_date = str(latest_date + timedelta(days=1)).split(" ")[0]
    else:
        print("No Monzo tag found in the file.")

    
    
    return latest_date

