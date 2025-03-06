import logging
import os
import numpy as np
import pandas as pd
from finance import constants
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


# File paths
CSV_FILE = constants.CSV_FILE
EXCEL_FILE = constants.EXCEL_FILE
SHEET_NAME = constants.SHEET_NAME


# Function to append CSV data to the end of an existing table in Excel
def append_csv_to_excel(csv_file=CSV_FILE, excel_file=EXCEL_FILE, sheet_name=SHEET_NAME):
    
    # Ensure the monzo folder exists
    os.makedirs("monzo", exist_ok=True)
    
    # Check if CSV file exists
    if not os.path.exists(csv_file):
        print(f"Error: CSV file '{csv_file}' not found.")
        return
    
    # Read the CSV file
    df = pd.read_csv(csv_file)

    # Select the relevant columns
    selected_columns = ["Date", "Amount (GBP)", "Merchant", "Category"]
    df_filtered = df[selected_columns] 
    df_filtered["Type"] = None
    df_filtered["Account"] = "Monzo"
    df_filtered["Balance"] = '=SUMPRODUCT([Amount],--([Date]<=[@Date]), (([Type]="Expenses") + ([Type]="Savings")) * (-1) + ([Type] = "Income"))'
    df_filtered["Effective Date"] = '=IF(AND([@Type]="Income", shift_income_status = "Active", DAY([@Date])>=shift_income_starting_date),DATE(YEAR([@Date]),MONTH([@Date])+1,1),([@Date]))'

    # format responses
    df_filtered["Date"] = pd.to_datetime(df_filtered["Date"], errors='coerce')
    df_filtered["Type"] = np.where(df_filtered["Amount (GBP)"] < 0, "Expenses", "Income")
    df_filtered["Amount (GBP)"] = df_filtered["Amount (GBP)"].abs() 

    # Update 'Category' with custom mappings
    misc_mapping = "Misc / Unknown"
    category_mapping = {
        "eating_out": "Food & Eating Out",
        "cash": misc_mapping,
        "other": misc_mapping
    }
    df_filtered["Category"] = df_filtered["Category"].replace(category_mapping).str.title()   

    logging.debug(df_filtered)

    # Check if Excel file exists else create one 
    if not os.path.exists(excel_file):
        df.to_excel(excel_file, sheet_name=sheet_name, index=False, engine="openpyxl")
        print(f"Created new Excel file: {excel_file}")
        return

    # Load existing workbook and find the last row
    workbook = load_workbook(excel_file)
    sheet = workbook[sheet_name]

    # Get the table
    table = sheet.tables['Tracking']

    # Get the current table range
    start_cell, end_cell = table.ref.split(':')
    start_col_letter = start_cell[0]
    end_col_letter = end_cell[0]
    end_row = int(end_cell[1:])

    # Table Formatting
    font_style = Font(size=10)
    indent_style = Alignment(indent=1) 
    indent_style_small = Alignment(indent=0.5) 
    indent_style_left = Alignment(indent = 1, horizontal="left")

    # Loop through all rows in df_filtered and add each to the table
    for i, (_, row) in enumerate(df_filtered.iterrows()):
        new_row_index = end_row + 1 + i

        # Column C (Date)
        cell = sheet.cell(row=new_row_index, column=3, value=row["Date"])
        cell.font = font_style
        cell.alignment = indent_style_left
        cell.number_format = "DD-MMM-YY"  # This applies the date format in Excel


        # Column D (Type)
        cell = sheet.cell(row=new_row_index, column=4, value=row["Type"])
        cell.font = font_style

        # Column E (Category)
        cell = sheet.cell(row=new_row_index, column=5, value=row["Category"])
        cell.font = font_style
        cell.alignment = indent_style

        # Column F (Amount)
        cell = sheet.cell(row=new_row_index, column=6, value=row["Amount (GBP)"])
        cell.font = font_style
        cell.alignment = indent_style_left

        # Column G (Merchant)
        cell = sheet.cell(row=new_row_index, column=7, value=row["Merchant"])
        cell.font = font_style
        cell.alignment = indent_style

        # Column H (Balance)
        row["Balance"] = f'=SUMPRODUCT([Amount],--([Date]<=C{new_row_index}), (([Type]="Expenses") + ([Type]="Savings")) * (-1) + ([Type] = "Income"))'
        cell = sheet.cell(row=new_row_index, column=8, value=row["Balance"])
        cell.font = font_style
        cell.alignment = indent_style

        # Column I (Account)
        cell = sheet.cell(row=new_row_index, column=9, value=row["Account"])
        cell.font = font_style
        cell.alignment = indent_style

        # Column J (Effective Date)
        row['Effective Date'] = f'=IF(AND(D{new_row_index}="Income", shift_income_status = "Active", DAY(C{new_row_index})>=shift_income_starting_date),DATE(YEAR(C{new_row_index}),MONTH(C{new_row_index})+1,1),(C{new_row_index}))'
        cell = sheet.cell(row=new_row_index, column=10, value=row["Effective Date"])
        cell.font = font_style
        cell.alignment = indent_style

    # Update the table's range to include all new rows
    total_new_rows = len(df_filtered)
    new_end_row = end_row + total_new_rows
    table.ref = f"{start_col_letter}{start_cell[1:]}:{end_col_letter}{new_end_row}"
    logging.info(f"Data successfully appended to {excel_file} at row {end_row}")

    # Save the workbook
    workbook.save(excel_file)